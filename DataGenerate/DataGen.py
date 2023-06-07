import openpyxl

def dataGenerator():
    wb=openpyxl.load_workbook("C:/Users/amnessa/Desktop/Python courses/SeleniumDataSheet.xlsx")
    sh=wb['Sheet1']
    r=sh.max_row
    li=[]
    li1=[]
    for i in range(1,r+1):
        li1=[]
        un=sh.cell(i,1)
        ue=sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,ue.value)
        li.insert(i-1,li1)
    print(li)
    return li